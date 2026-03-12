import fitz  # PyMuPDF
import json
import pandas as pd
from pathlib import Path
from typing import Dict, List, Any, Optional
import re
import os
from dotenv import load_dotenv
import base64
import glob
import pymupdf4llm
from io import StringIO
from datetime import datetime
import uuid
from edgar_parser import parse_markdown_content, apply_excel_formatting
from document_metadata import DocumentMetadata, MAX_TOKENS
from langchain_deepseek import ChatDeepSeek
from langchain.prompts import ChatPromptTemplate
import sys
from openpyxl.utils import get_column_letter
# Try to import the prompt templates, handle both ways of running the script
try:
    from backend.prompts.document_metadata_prompt import document_metadata_prompt_template
except ModuleNotFoundError:
    # Add project root to Python path if running as a script
    project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), "../.."))
    if project_root not in sys.path:
        sys.path.append(project_root)
    from backend.prompts.document_metadata_prompt import document_metadata_prompt_template

from edgar_parser_tools import (
    clean_edgar_table_perc_currency,
    clean_dataframe_table,
    prepare_table_with_headers,
    get_char_frequency,
    calculate_frequency_similarity,
    find_best_matching_table
)
# Load environment variables from .env.local
load_dotenv(".env.local")

DATA_ROOT = os.environ.get("DATA_ROOT")

TICKER_LIST = [
    'RBLX'
]

SOURCE_CATALOG_LIST = [
    # "earnings_call_transcript",
    # "company_presentation",
    # "company_presentation_slides",
    # "earnings_slides",

    "research-reports",
    # "research-ms",
    # "research-bcly",
    # "research-rbc",
    # "research-ubs",
    # "research-boa",
    # "research-db",
    # "research-wfgo",
    # "research-jfr",
    # "research-nham",
    # "research-bmo",
    # "research-roth",
    # "research-wbsh",
    # "research-jpm",
    # "research-opco",
    # "research-ppr",
    # "research-sig",
    # "research-tdc",
    # "expert-as"
]

BATCH_SIZE = 5  # Reduced from 5 to prevent memory issues and API rate limiting


class PDFParserFitz:
    """
    PDF parser for biopharma financial documents using PyMuPDF (Fitz)
    Following edgar_parser.py design patterns
    """
    
    def __init__(self, pdf_path: str, earnings_cal: Optional[Dict[str, Any]] = None, source_catalog: str = None):
        self.pdf_path = pdf_path
        self.doc = fitz.open(pdf_path)
        self.path_obj = Path(pdf_path)
        self.filename_stem = self.path_obj.stem
        self.uuid = str(uuid.uuid4())
        
        # Create a directory for outputs related to this PDF, inside the PDF's directory
        self.output_dir = self.path_obj.parent / self.filename_stem
        self.output_dir.mkdir(exist_ok=True)

        self.parsed_data = []
        self.tables_data = []
        self.src_metadata: Optional[DocumentMetadata] = None
        self.publish_date = ""
        self.report_date = ""
        self.year_q_no = ""
        self.earnings_cal = earnings_cal
        self.source_catalog = source_catalog
    
    def extract_table_summary(self, table_content: str) -> str:
        """Extract table summary from table content"""
        return table_content
    
    def parse_document(self) -> Dict[str, Any]:
        """Main parsing function - follows edgar_parser pattern"""
        for page_num in range(len(self.doc)):
            page_data = self.parse_page(page_num)
            self.parsed_data.extend(page_data)
        

        for element in self.parsed_data:
            if element["type"] == "table":
                element["summary"] = self.extract_table_summary(element["content"])

        self.save_outputs()
        return self.get_summary()

    def parse_page(self, page_num: int) -> List[Dict]:
        """Parse a single page - main processing function"""
        page_number = page_num + 1
        md_file_path = self.output_dir / f"page{page_number:03d}_image_001.md"
        
        # Check if .md file exists for this page
        if md_file_path.exists():
            # Load content from existing .md file
            try:
                with open(md_file_path, 'r', encoding='utf-8') as f:
                    md_content = f.read()
                
                # Still need to load the page for visual elements and consistent processing
                page = self.doc.load_page(page_num)
                
                # Extract text elements from .md content
                text_elements = self.extract_text_with_layout(page, page_num, md_content)
                return text_elements
            except Exception as e:
                print(f"Error reading .md file {md_file_path}: {e}")
                print("Falling back to original parsing method...")
                # Fall through to original parsing
        
        # Original parsing method when .md file doesn't exist
        page = self.doc.load_page(page_num)
        page_elements = []

        # First pass: get elements with bounding boxes
        visual_elements = self.extract_images_and_charts(page, page_num)

        # Combine and sort elements with bboxes by their vertical position
        other_elements = visual_elements
        
        
        # Second pass: extract text and titles
        text_elements = self.extract_text_with_layout(page, page_num)

        # Merge all elements in order. Since text_elements don't have bbox,
        # and pymupdf4llm returns them in reading order, we can assume they
        # are in a single block for layout purposes. We'll treat them as
        # being at the top of the page for now. A more sophisticated approach
        # might be needed if text is heavily interspersed with tables/images.
        page_elements.extend(text_elements)
        page_elements.extend(other_elements)
        
        return page_elements

    def extract_text_with_layout(self, page: fitz.Page, page_num: int, md_content: str = None) -> List[Dict]:
        """Extract text while preserving layout structure, and detecting titles using PyMuPDF4LLM."""
        elements = []
        
        if md_content is None:
            try:
                md_content = pymupdf4llm.to_markdown(self.doc, pages=[page_num])
            except Exception as e:
                print(f"    Failed to extract text from page {page_num + 1}: {e}")
                return elements  # Return empty elements if text extraction fails
        
        if md_content is None:
            print(f"    No text content extracted from page {page_num + 1}")
            return elements

        lines = md_content.split('\n')
        text_buffer = []
        table_buffer = []

        for line in lines:
            stripped_line = line.strip()
            if not stripped_line:
                # Preserve empty lines as "\n" in text buffer
                text_buffer.append("\n")

                # flush table buffer
                if table_buffer:
                    content = "\n".join(table_buffer).strip()
                    if content:
                        elements.append({
                            "type": "table",
                            "summary": "",
                            "content": content,
                            "page_no": page_num + 1,
                        })
                    table_buffer = []

                continue

            if stripped_line.startswith('#'):
                # Flush text buffer if it contains content
                if text_buffer:
                    content = "\n".join(text_buffer).strip()
                    if content:
                        elements.append({
                            "type": "text",
                            "content": content,
                            "page_no": page_num + 1,
                        })
                    text_buffer = []

                if table_buffer:
                    content = "\n".join(table_buffer).strip()
                    if content:
                        elements.append({
                            "type": "table",
                            "summary": "",
                            "content": content,
                            "page_no": page_num + 1,
                        })
                    table_buffer = []

                # Process the title
                level = 0
                for char in stripped_line:
                    if char == '#':
                        level += 1
                    else:
                        break
                
                title = stripped_line[level:].strip()
                
                if title:
                    elements.append({
                        "type": "title",
                        "level": level - 1,  # Convert to 0-based indexing
                        "content": title,
                        "page_no": page_num + 1,
                    })
            elif stripped_line.startswith("|"):
                # Flush text buffer if it contains content
                if text_buffer:
                    content = "\n".join(text_buffer).strip()
                    if content:
                        elements.append({
                            "type": "text",
                            "content": content,
                            "page_no": page_num + 1,
                        })
                    text_buffer = []

                # Process the table 
                table_buffer.append(stripped_line)
            else:
                # Flush table buffer before processing text content
                if table_buffer:
                    content = "\n".join(table_buffer).strip()
                    if content:
                        elements.append({
                            "type": "table",
                            "summary": "",
                            "content": content,
                            "page_no": page_num + 1,
                        })
                    table_buffer = []
                
                text_buffer.append(stripped_line)

        # Flush any remaining text in the buffer
        if text_buffer:
            content = "\n".join(text_buffer).strip()
            if content:
                elements.append({
                    "type": "text",
                    "content": content,
                    "page_no": page_num + 1,
                })

        if table_buffer:
            content = "\n".join(table_buffer).strip()
            if content:
                elements.append({
                    "type": "table",
                    "summary": "",
                    "content": content,
                    "page_no": page_num + 1,
                })

        return elements


    def extract_images_and_charts(self, page: fitz.Page, page_num: int) -> List[Dict]:
        """Looks for previously extracted images for the page, encodes them, and returns image elements."""
        elements_found = []
        
        page_prefix = f"page{page_num+1:03d}"
        # Using glob to find files matching the pattern
        search_pattern = self.output_dir / f"{page_prefix}_*.png"
        
        # In Python 3.9+ Path objects can be passed to glob
        # For broader compatibility, convert to string
        image_files = sorted(glob.glob(str(search_pattern)))

        for image_path_str in image_files:
            image_path = Path(image_path_str)
            image_filename = image_path.name
            # The image_id is derived from the stem of the image_path, which is the filename without its suffix.
            image_id = image_path.stem
            
            try:
                with open(image_path, "rb") as f:
                    image_bytes = f.read()
                
                base64_content = base64.b64encode(image_bytes).decode('utf-8')

                # check if image_filename.md or image_filename_qwen.md exists
                # load to markdown_content
                markdown_content = ""
                md_path_plain = self.output_dir / f"{image_id}.md"

                if md_path_plain.exists():
                    with open(md_path_plain, "r", encoding='utf-8') as f:
                        markdown_content = f.read()
                
                elements_found.append({
                    "type": "image",
                    "image_name": image_filename,
                    "content": markdown_content,
                    "image": base64_content,
                    "page_no": page_num + 1,
                    "image_id": image_id,
                    "bbox": (0, 0, 0, 0) # Placeholder as bbox is not available from file
                })

            except Exception as e:
                print(f"Error processing image file {image_filename}: {e}")

        return elements_found

    def extract_src_metadata(self):
        """Extracts source metadata from the markdown file."""
        content = ''
        image_content = ''

        md_path = self.output_dir / f"{self.filename_stem}.md"
        if not md_path.exists():
            print(f"Markdown file not found at {md_path}, skipping metadata extraction.")
            return

        with open(md_path, "r", encoding="utf-8") as f:
            content = f.read()

        # image_md_path = self.output_dir / f"page001_image_001.md"
        # if image_md_path.exists():
        #     with open(image_md_path, "r", encoding="utf-8") as f:
        #         image_content = f.read()

        document_content = content
        if len(document_content) > MAX_TOKENS * 4:
            print(f"Document is too long, truncate the document to {MAX_TOKENS} tokens")
            index = int(0.7 * MAX_TOKENS * 4)
            document_content = content[:index] + "...\n\n...\n"
            index = int(0.3 * MAX_TOKENS * 4)
            document_content += content[-index:]

        filename_string = f"## Filename: {self.filename_stem}.pdf"
        document_content = filename_string + "\n\n" + document_content
        
        # Add earnings calendar data if available
        calendar_prompt = ""
        if self.earnings_cal:
            calendar_prompt = "<earnings_calendar>\n"
            calendar_prompt += "Format: YYYY-MM-DD [Fiscal Quarter]\n"
            
            # Serialize earnings calendar data
            records = self.earnings_cal.get("tipranks", [])
            for record in records:  # Limit to first 10 records to avoid token limits
                if 'Report Date' in record:
                    fiscal_quarter = record.get('Fiscal Quarter', 'N/A')
                    if fiscal_quarter and fiscal_quarter != 'N/A':
                        # Clean up fiscal quarter format
                        fiscal_quarter = fiscal_quarter.replace("q", "Q")
                        fiscal_quarter = fiscal_quarter.replace(" ", "")
                        fiscal_quarter = fiscal_quarter.replace("(", "")
                        fiscal_quarter = fiscal_quarter.replace(")", "")
                        fiscal_quarter = fiscal_quarter.replace("-", "")
                        
                        calendar_prompt += f"{record['Report Date']} [{fiscal_quarter}]\n"
            calendar_prompt += "</earnings_calendar>\n"
            
        
        try:
            deepseek_api_key = os.getenv("DEEPSEEK_API_KEY")
            if not deepseek_api_key:
                raise ValueError("DEEPSEEK_API_KEY environment variable not set")

            llm = ChatDeepSeek(
                model="deepseek-chat",
                temperature=0.7,
                api_key=deepseek_api_key
            )
            structured_llm = llm.with_structured_output(DocumentMetadata)
            prompt_template_obj = ChatPromptTemplate.from_template(document_metadata_prompt_template)
            prompt = prompt_template_obj.format(content=document_content, earnings_calendar=calendar_prompt)

            response = structured_llm.invoke(prompt)
            self.src_metadata = response

            if self.source_catalog != None:
                # Override form_type and source_institution for specific source catalogs
                self.src_metadata.form_type = self.source_catalog
                self.src_metadata.source_institution = "issuer"

            self.publish_date, self.report_date = self._get_early_and_late_dates()
            self.year_q_no = self._calculate_year_quarter()

            # output the response (src_metadata) along with self.filename_stem, publish_date and report_date, year_q_no, to a json file
            metadata_path = self.output_dir / f"src_metadata.json"
            with open(metadata_path, "w", encoding="utf-8") as f:
                metadata_output = {
                    "filename": self.filename_stem+'.pdf',
                    "uuid": self.uuid,
                    "prompt_template": document_metadata_prompt_template,
                    "content": document_content,
                    "prompt": prompt,
                    "publish_date": self.publish_date,
                    "report_date": self.report_date,
                    "year_q_no": self.year_q_no,
                    "src_metadata": self.src_metadata.model_dump()
                }
                json.dump(metadata_output, f, indent=4, ensure_ascii=False)

            print(f"Successfully extracted metadata for {self.filename_stem}")

        except Exception as e:
            print(f"Error extracting source metadata for {self.filename_stem}: {e}")

    def save_outputs(self):
        """Save all output files"""
        self.save_markdown_output()
        self.extract_src_metadata()
        self.save_tables_output()
        self.save_images_tables_output()
        self.save_json_output()

    def save_json_output(self):
        """Save parsed data as JSON file"""
        output_path = self.output_dir / f"{self.filename_stem}.json"
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(self.parsed_data, f, indent=4, ensure_ascii=False)

    def _table_meets_threshold(self, markdown_table: str) -> bool:
        """
        Check if a markdown table meets the threshold (more than 1 data row).
        
        Args:
            markdown_table: Markdown table string
            
        Returns:
            True if table has more than 1 data row, False otherwise
        """
        try:
            lines = [l.strip() for l in markdown_table.strip().split('\n') if l.strip()]
            
            # Need at least 3 lines: header, separator, and 2+ data rows
            # But separator might be removed, so check for at least 3 lines total
            # If there's a separator (line with only |-:), skip it
            data_row_count = 0
            for i, line in enumerate(lines):
                if i == 0:
                    # First line is header
                    continue
                # Check if it's a separator line (contains only |, -, :, and spaces)
                if all(c in '-|: ' for c in line):
                    # This is the separator, skip it
                    continue
                # This is a data row
                data_row_count += 1
            
            # Threshold: more than 1 data row
            return data_row_count > 1
        except Exception as e:
            # If we can't parse it, assume it meets threshold to be safe
            print(f"Warning: Could not check table threshold: {e}")
            return True

    def save_markdown_output(self):
        """Generate markdown file with readable text content"""
        output_path = self.output_dir / f"{self.filename_stem}.md"
        markdown_content = []

        current_page = 1
        table_number = 1
        
        for element in self.parsed_data:
            if element["page_no"] > current_page:
                for index in range(current_page, element["page_no"]):
                    if index == 1:
                        markdown_content.append(f"\n\n---\n")
                    else:
                        markdown_content.append(f"\n{index}\n\n---\n")
                current_page = element["page_no"]

            if element["type"] == "text":
                markdown_content.append(element["content"])
            elif element["type"] == "title":
                head_string = "#" * element["level"] + "#"
                markdown_content.append(f"\n{head_string} {element['content']}\n")
            elif element["type"] == "table":
                # content is already in markdown format
                # Check if table meets threshold (more than 1 data row)
                table_meets_threshold = self._table_meets_threshold(element["content"])
                
                if table_meets_threshold:
                    # Output with table number for valid tables
                    markdown_content.append(f"[[Table{table_number}]]\n")
                    markdown_content.append(element["content"]+"\n")
                    markdown_content.append(f"[/Table{table_number}]\n\n")
                    table_number += 1
                else:
                    # Output without table number for small tables
                    markdown_content.append(f"[[Table]]\n")
                    markdown_content.append(element["content"]+"\n")
                    markdown_content.append(f"[/Table]\n\n")
            elif element["type"] == "image":
                markdown_content.append(f"\n![image]({element['image_name']})\n")
                if element.get('content'):
                    markdown_content.append(element['content'])
                    image_basename = Path(element['image_name']).stem
                    markdown_content.append(f"\n[[{image_basename}.md]]\n")
                    markdown_content.append(f"[[{image_basename}.xlsx]]\n")
                    markdown_content.append(f"[/IMG]\n")
        
        with open(output_path, "w", encoding="utf-8") as f:
            f.write("\n".join(markdown_content))




    def create_content_table(self, df_tables, form_type="Research Report"):
        """
        Create a content table with SEC metadata and table summaries
        
        Args:
            df_tables: List of pandas DataFrames containing the tables
            form_type: Type of SEC form ('10-K', '10-Q', '8-K')
            
        Returns:
            Tuple of (markdown_content_table, content_dataframe)
        """
        import pandas as pd
        from datetime import datetime
        
        # Prepare SEC metadata section
        metadata_info = []
        
        if self.src_metadata:
            if self.src_metadata.primary_ticker:
                metadata_info.append(["Ticker", self.src_metadata.primary_ticker])
            if self.src_metadata.highlights:
                metadata_info.append(["Highlights", " • ".join(self.src_metadata.highlights)])
            if self.src_metadata.primary_doc_description:
                metadata_info.append(["Report Type", self.src_metadata.primary_doc_description])
            
            if self.publish_date:
                metadata_info.append(["Publish Date", self.publish_date])
            if self.report_date:
                metadata_info.append(["Report Date", self.report_date])

            # calculate year_q_no
            if self.year_q_no:
                metadata_info.append(["Year and Quarter", self.year_q_no])

            if self.src_metadata.lead_analysts:
                metadata_info.append(["Lead Analysts", ", ".join(self.src_metadata.lead_analysts)])

            metadata_info.append(["Filename", self.filename_stem+'.pdf'])





        else:
            metadata_info.extend([
                ["Report Type", form_type],
                ["Processing Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                ["Note", "Metadata not available"]
            ])
        
        # Add separator
        metadata_info.append(["---", "---"])
        
        # Prepare table summaries
        table_summaries = []
        if df_tables:
            for i, table in enumerate(df_tables, 1):
                if not table.empty and len(table.columns) > 0:
                    # Get first column name and sample values
                    first_col_name = str(table.columns[0])
                    
                    # Get non-empty values from first column (limit to first 10)
                    first_col_values = []
                    for val in table.iloc[:, 0].dropna().astype(str):
                        if val.strip() and val.strip() != 'nan':
                            first_col_values.append(val.strip())
                    
                    # Create summary text
                    if first_col_values:
                        summary_text = f"Items: {', '.join(first_col_values[:5])}"
                        if len(first_col_values) > 5:
                            summary_text += f" ... (+{len(first_col_values)-5} more)"
                    else:
                        summary_text = f"No item available"
                    
                    table_summaries.append([f"Table {i}", summary_text])
                else:
                    table_summaries.append([f"Table {i}", "Empty table"])
        else:
            table_summaries.append(["Notice", "No PDF tables found, check the image tables."])
        
        # Combine all content
        all_content = metadata_info + table_summaries
        
        # Create DataFrame for Excel output
        content_df = pd.DataFrame(all_content, columns=["Field", "Value"])
        
        # Create markdown table
        markdown_lines = ["# Overview", "", ""]
        markdown_lines.append("## Filing Information")
        markdown_lines.append("")
        
        # Add metadata as markdown table
        markdown_lines.append("| Field | Value |")
        markdown_lines.append("|-------|-------|")
        
        for field, value in metadata_info:
            if field != "---":  # Skip separator row
                markdown_lines.append(f"| {field} | {value} |")
        
        markdown_lines.extend(["", "## Table Summary", ""])
        markdown_lines.append("| Table | First Column Items |")
        markdown_lines.append("|-------|--------------------|")
        
        for table_name, summary in table_summaries:
            # Escape pipe characters in summary text
            summary_escaped = summary.replace("|", "\\|")
            markdown_lines.append(f"| {table_name} | {summary_escaped} |")
        
        markdown_lines.extend(["", "---", "", ""])
        
        markdown_content = "\n".join(markdown_lines)
        
        return markdown_content, content_df

    def _get_early_and_late_dates(self):
        """
        Calculates the earliest and latest dates from metadata attributes.

        This function iterates through a predefined list of date-related attributes 
        of the `src_metadata` object, attempts to parse any valid date strings, 
        and identifies the minimum (earliest) and maximum (latest) dates among them.
        Invalid or missing date values are ignored.

        Returns:
            If no valid dates are found, returns (None, None).
        """
        dates = []
        date_fields = ['report_date', 'filing_date', 'interview_date', 'publish_date']
        for field in date_fields:
            date_str = getattr(self.src_metadata, field, None)
            if date_str and isinstance(date_str, str):
                try:
                    # Attempt to parse date in "YYYY-MM-DD" format
                    parsed_date = datetime.strptime(date_str, "%Y-%m-%d")
                    dates.append(parsed_date)
                except ValueError:
                    # Ignore if the date string is not in the expected format
                    continue
        
        if not dates:
            return None, None
        
        self.publish_date = max(dates).strftime("%Y-%m-%d")
        self.report_date = min(dates).strftime("%Y-%m-%d")
        return self.publish_date, self.report_date



    def _calculate_year_quarter(self) -> str:
        """
        Calculate year and quarter from filing date.
        
        Args:
            filing_date: Date string in format "YYYY-MM-DD"
            
        Returns:
            String in format "YYYYQX" (e.g., "2024Q3")
        """

        if self.report_date:
            date_obj = datetime.strptime(self.report_date, "%Y-%m-%d")
        elif self.publish_date:
            date_obj = datetime.strptime(self.publish_date, "%Y-%m-%d")
        else:
            return ""

        year = date_obj.year
        month = date_obj.month
        
        if month <= 3:
            quarter = 1
        elif month <= 6:
            quarter = 2
        elif month <= 9:
            quarter = 3
        else:
            quarter = 4
        
        self.year_q_no = f"{year}Q{quarter}"
        return self.year_q_no
        




    def create_image_analysis_excel(self, image_filename, image_md_content, form_type="Research Report"):
        """
        Create Excel file from image markdown content with multiple sheets
        
        Args:
            image_filename: Name of the image file
            image_md_content: Markdown content to parse
            sec_metadata: SEC filing metadata
            form_type: Form type (10-K, 10-Q, 8-K)
            output_path: Path where to save the Excel file
        """
        import pandas as pd
        from datetime import datetime
        
        # Parse markdown content
        sections, text_content = parse_markdown_content(image_md_content)
        
        # Create content overview similar to main filing
        metadata_info = []
    
        if self.src_metadata:
            if self.src_metadata.primary_ticker:
                metadata_info.append(["Ticker", self.src_metadata.primary_ticker])
            if self.src_metadata.highlights:
                metadata_info.append(["Highlights", " • ".join(self.src_metadata.highlights)])
            if self.src_metadata.primary_doc_description:
                metadata_info.append(["Report Type", self.src_metadata.primary_doc_description])
            
            if self.publish_date:
                metadata_info.append(["Publish Date", self.publish_date])
            if self.report_date:
                metadata_info.append(["Report Date", self.report_date])

            # calculate year_q_no
            if self.year_q_no:
                metadata_info.append(["Year and Quarter", self.year_q_no])

            if self.src_metadata.lead_analysts:
                metadata_info.append(["Lead Analysts", ", ".join(self.src_metadata.lead_analysts)])

            metadata_info.append(["Image Filename", image_filename])

        else:
            metadata_info.extend([
                ["Report Type", form_type],
                ["Processing Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                ["Image Filename", image_filename],
                ["Note", "Metadata not available"]
            ])
        
        # Add separator
        metadata_info.append(["---", "---"])
        
        # Add table summaries
        table_counter = 1
        for section in sections:
            for table_info in section.get('tables', []):
                df = table_info['dataframe']
                if not df.empty and len(df.columns) > 0:
                    # Get first column name and sample values (same logic as main filing)
                    first_col_name = str(df.columns[0])
                    
                    # Get non-empty values from first column (limit to first 10)
                    first_col_values = []
                    for val in df.iloc[:, 0].dropna().astype(str):
                        if val.strip() and val.strip() != 'nan':
                            first_col_values.append(val.strip())
                        if len(first_col_values) >= 10:  # Limit to prevent overly long lists
                            break
                    
                    # Create summary text
                    if first_col_values:
                        summary_text = f"Items: {', '.join(first_col_values[:5])}"
                        if len(first_col_values) > 5:
                            summary_text += f" ... (+{len(first_col_values)-5} more)"
                    else:
                        summary_text = f""
                    
                    metadata_info.append([f"Table {table_counter}", summary_text])
                else:
                    metadata_info.append([f"Table {table_counter}", ""])
                table_counter += 1
        
        # Add text content at the end
        if text_content.strip():
            metadata_info.extend([
                ["---", "---"],
                ["Contextual Info", text_content]
            ])
        
        # Create content overview DataFrame
        content_df = pd.DataFrame(metadata_info, columns=["Field", "Value"])

        excel_filename = image_filename.rsplit('.', 1)[0] + '.xlsx'
        
        # Create Excel file with multiple sheets and formatting
        with pd.ExcelWriter(self.output_dir / excel_filename, engine='openpyxl') as writer:
            # Write content overview as first sheet
            content_df.to_excel(
                writer,
                sheet_name="Overview",
                index=False,
                header=True
            )
                
            # Set column widths for Overview sheet
            worksheet = writer.sheets["Overview"]
            pixel_width = 256
            # This is an approximation. The actual width depends on the default font.
            char_width = pixel_width / 7
            for col_idx in range(1, worksheet.max_column + 1):
                worksheet.column_dimensions[get_column_letter(col_idx)].width = char_width
            
            # Write each table as separate sheet with formatting
            table_counter = 1
            for section in sections:
                section_title = section.get('title', f'Section {table_counter}')
                
                for table_info in section.get('tables', []):
                    df = table_info['dataframe']
                    if not df.empty:
                        # Clean sheet name (Excel has restrictions)
                        sheet_name = f"Table {table_counter}"
                        if len(section_title) > 0:
                            clean_title = re.sub(r'[^\w\s-]', '', section_title)[:20]  # Remove special chars, limit length
                            if clean_title.strip():
                                sheet_name = f"T{table_counter}_{clean_title.strip().replace(' ', '_')}"
                        
                        # Ensure sheet name is valid for Excel
                        sheet_name = sheet_name[:31]  # Excel limit
                        
                        # Write DataFrame to sheet
                        df.to_excel(
                            writer,
                            sheet_name=sheet_name,
                            index=False,
                            header=True
                        )
                            
                        # Set column widths for table sheet
                        worksheet = writer.sheets[sheet_name]
                        for col_idx in range(1, worksheet.max_column + 1):
                            worksheet.column_dimensions[get_column_letter(col_idx)].width = char_width
                        
                        # Apply formatting to the worksheet
                        worksheet = writer.sheets[sheet_name]
                        
                        # Apply header formatting
                        header_formatting = table_info.get('header_formatting', [])
                        for col_idx, (header, fmt) in enumerate(zip(df.columns, header_formatting)):
                            apply_excel_formatting(worksheet, 1, col_idx + 1, header, fmt)
                        
                        # Apply data formatting
                        data_formatting = table_info.get('data_formatting', [])
                        for row_idx, (row_data, row_fmt) in enumerate(zip(df.values, data_formatting)):
                            for col_idx, (cell_value, cell_fmt) in enumerate(zip(row_data, row_fmt)):
                                apply_excel_formatting(worksheet, row_idx + 2, col_idx + 1, cell_value, cell_fmt)
                        
                        table_counter += 1
            
            # If no tables found, add a summary sheet with text content
            if table_counter == 1:
                text_content = []
                for section in sections:
                    if section.get('title'):
                        text_content.append([section['title'], 'Header'])
                    for line in section.get('content', []):
                        if line.strip():
                            text_content.append([line.strip(), 'Text'])
                
                if text_content:
                    text_df = pd.DataFrame(text_content, columns=["Content", "Type"])
                    text_df.to_excel(
                        writer,
                        sheet_name="Text Content",
                        index=False,
                        header=True
                    )
        
                    # Set column widths for Text Content sheet
                    worksheet = writer.sheets["Text Content"]
                    for col_idx in range(1, worksheet.max_column + 1):
                        worksheet.column_dimensions[get_column_letter(col_idx)].width = char_width
    
        print(f"Created image analysis Excel file with formatting: {self.output_dir / excel_filename}")






    def save_tables_output(self):
        df_tables = []
        table_index = 0
        print("\n--- Parsing Markdown Tables into DataFrames ---")
        for i, element in enumerate(self.parsed_data):
            if element["type"] != "table":
                continue
            
            table_index += 1

            markdown_table = element['content']
            page_num = element['page_no']
            print(f"\n--- Processing Table {i+1} on Page {page_num} ---")
            print("Original Markdown Table:")
            print(markdown_table)

            try:
                # Prepare markdown for pandas parsing
                lines = [l.strip() for l in markdown_table.strip().split('\n')]
                
                # remove separator line
                if len(lines) > 1 and all(c in '-|: ' for c in lines[1]):
                    del lines[1]
                
                # remove leading/trailing pipes and create a CSV-like string
                cleaned_lines = [l.strip()[1:-1].strip() if l.startswith('|') and l.endswith('|') else l for l in lines]
                
                # Important: Tables with || at start have empty first column for row labels
                # Example: || 2020A | 2021A | means ["", "2020A", "2021A"] 
                # The empty first column should contain row names like "Revenue", "Gross Margin"
                # We preserve this structure - pandas will use empty string as first column name
                
                # Create summary WITHOUT table number for now - will be added later after filtering
                first_col_values = []

                # Skip header row and extract first column values from the rest
                if len(cleaned_lines) > 1:
                    for line in cleaned_lines[1:]:
                        col_value = line.split('|')[0].strip()
                        if col_value and col_value.lower() != 'nan':
                            first_col_values.append(col_value)
                
                if first_col_values:
                    summary_text = f"Items: {', '.join(first_col_values[:5])}"
                    if len(first_col_values) > 5:
                        summary_text += f" ... (+{len(first_col_values)-5} more)"
                else:
                    summary_text = ""
                
                # Store summary without table number - will be added after filtering
                element['summary'] = summary_text
                self.parsed_data[i]['summary'] = summary_text

                csv_string = "\n".join(cleaned_lines)
                table_io = StringIO(csv_string)
                
                # Explicitly set first row as header (row 0)
                df = pd.read_csv(table_io, sep='|', skipinitialspace=True, header=0)
                
                # clean column names
                df.columns = df.columns.str.strip()

                print("\nGenerated DataFrame:")
                print(f"Columns: {list(df.columns)}")
                print(df.to_string())
                df_tables.append(df)
            except Exception as e:
                print(f"!!! Could not parse table {i+1} into DataFrame. Error: {e}")
                print("    Markdown content was:\n" + markdown_table)
        
        print("\n--- Finished parsing tables. ---")

        markdown_filename = self.output_dir / f"{self.filename_stem}_[tables].md"
        excel_filename = self.output_dir / f"{self.filename_stem}.xlsx"



        # Mark which tables meet the threshold (more than 1 row)
        table_validity = []
        valid_tables_only = []
        if df_tables:
            for i, table in enumerate(df_tables):
                is_valid = len(table) > 1  # More than 1 data row
                table_validity.append(is_valid)
                if is_valid:
                    valid_tables_only.append(table)
                else:
                    print(f"📝 DEBUG: Table {i+1} is small - only has {len(table)} row(s)")
            
            valid_count = len(valid_tables_only)
            small_count = len(df_tables) - valid_count
            if small_count > 0:
                print(f"🗑️ DEBUG: Found {small_count} small tables. {valid_count} valid tables will be numbered.")
            else:
                print(f"✅ DEBUG: All {len(df_tables)} tables are valid, no filtering needed.")

        # Create content table with metadata and table summaries (only for valid tables)
        content_markdown, content_df = self.create_content_table(valid_tables_only, form_type="Research Report")

        markdown_perfect_tables = []

        # for every table in df_tables, output to a .md file with each table in a seperate page by "--- Table no.\n\n"
        if df_tables:  # Only create markdown tables file if tables were found
            with open(markdown_filename, "w", encoding="utf-8") as file:
                # Write content table first
                file.write(content_markdown)
                
                valid_table_counter = 0
                for i, table in enumerate(df_tables):
                    # Note: Don't call prepare_table_with_headers() here because
                    # pd.read_csv(header=0) already correctly extracted headers from markdown
                    table_string = table.to_markdown(index=False)

                    # Remove the header separator line entirely to save tokens
                    table_string_lines = table_string.split("\n")
                    second_line = table_string_lines[1]
                    while "--" in second_line:
                        second_line = second_line.replace("--", "-")

                    table_string_lines[1] = second_line
                    table_string = "\n".join(table_string_lines)
                    
                    while "  " in table_string:
                        table_string = table_string.replace("  ", " ")

                    # Only number valid tables
                    if table_validity[i]:
                        valid_table_counter += 1
                        markdown_perfect_tables.append(table_string)
                        # Output to markdown with table number
                        file.write(f"[[Table{valid_table_counter}]]\n")
                        file.write(table_string)
                        file.write(f"\n\n[/Table{valid_table_counter}]\n\n")
                    else:
                        # Output small table without number
                        file.write(f"[[Table]]\n")
                        file.write(table_string)
                        file.write(f"\n\n[/Table]\n\n")
        else:
            # Create file with just content table even if no tables found
            with open(markdown_filename, "w", encoding="utf-8") as file:
                file.write(content_markdown)

        # for every valid table, save it to one excel file with multiple sheets
        with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
            # Write content overview as first sheet
            content_df.to_excel(
                writer, 
                sheet_name="Overview", 
                index=False,
                header=True
            )
            
            # Set column widths for Overview sheet
            worksheet = writer.sheets["Overview"]
            pixel_width = 256
            # This is an approximation. The actual width depends on the default font.
            char_width = pixel_width / 7
            for col_idx in range(1, worksheet.max_column + 1):
                worksheet.column_dimensions[get_column_letter(col_idx)].width = char_width
            
            
            # Write individual valid tables only
            for i, table in enumerate(valid_tables_only):
                # Note: Don't call prepare_table_with_headers() here because
                # pd.read_csv(header=0) already correctly extracted headers from markdown
                # Calling prepare_table_with_headers would wrongly treat first data row as headers
                
                sheet_name = f"Table {i+1}"
                
                # Save to Excel with all cells as strings, no index
                table.to_excel(
                    writer, 
                    sheet_name=sheet_name, 
                    index=False,
                    header=True
                )
                # Set column widths for table sheet
                worksheet = writer.sheets[sheet_name]
                for col_idx in range(1, worksheet.max_column + 1):
                    worksheet.column_dimensions[get_column_letter(col_idx)].width = char_width




    def save_images_tables_output(self):
        """Save images and tables in separate markdown and Excel files"""
        # Images markdown
        image_elements = [e for e in self.parsed_data if e["type"] == "image"]
        if not image_elements:
            print("No images found in parsed data to save.")
            return
        
        for element in image_elements:
            if element.get('content'):
                self.create_image_analysis_excel(element["image_name"], element["content"])


    def dataframe_to_markdown(self, df: pd.DataFrame) -> str:
        """Convert DataFrame to markdown table"""
        if df.empty:
            return "Empty table"
        return df.to_markdown(index=False)

    def get_summary(self) -> Dict[str, Any]:
        """Return parsing summary statistics"""
        return {
            "total_pages": len(self.doc),
            "total_elements": len(self.parsed_data),
            "text_blocks": len([e for e in self.parsed_data if e["type"] == "text"]),
            "tables": len([e for e in self.parsed_data if e["type"] == "table"]),
            "images": len([e for e in self.parsed_data if e["type"] == "image"]),
            "charts": len([e for e in self.parsed_data if e["type"] == "chart"])
        }

def update_src_metadatas_entry(src_metadatas_path: str, uuid: str, updated_metadata: Dict[str, Any]) -> bool:
    """
    Update a specific entry in src_metadatas.json by UUID.
    
    Args:
        src_metadatas_path: Path to the src_metadatas.json file
        uuid: UUID of the document to update
        updated_metadata: New metadata to replace the existing entry
        
    Returns:
        True if updated successfully, False otherwise
    """
    try:
        # Load existing src_metadatas.json
        with open(src_metadatas_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        
        filings_by_form = data.get("filings_by_form", {})
        updated = False
        
        # Find and update the entry
        for form_type, filings in filings_by_form.items():
            for i, filing in enumerate(filings):
                if filing.get("uuid") == uuid:
                    filings[i] = updated_metadata
                    updated = True
                    print(f"Updated entry for UUID {uuid} in form_type {form_type}")
                    break
            if updated:
                break
        
        if updated:
            # Save the updated file
            with open(src_metadatas_path, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=4, ensure_ascii=False, default=str)
            return True
        else:
            print(f"No entry found with UUID {uuid}")
            return False
            
    except Exception as e:
        print(f"Error updating src_metadatas.json: {e}")
        return False


def find_document_by_uuid(src_metadatas_path: str, target_uuid: str) -> Optional[Dict[str, Any]]:
    """
    Find a document by its UUID in the src_metadatas.json file.
    
    Args:
        src_metadatas_path: Path to the src_metadatas.json file
        target_uuid: The UUID to search for
        
    Returns:
        The document metadata if found, None otherwise
    """
    try:
        with open(src_metadatas_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        
        filings_by_form = data.get("filings_by_form", {})
        
        for form_type, filings in filings_by_form.items():
            for filing in filings:
                if filing.get("uuid") == target_uuid:
                    return filing
        
        return None
    except Exception as e:
        print(f"Error loading src_metadatas.json: {e}")
        return None


def get_all_uuids(src_metadatas_path: str) -> List[str]:
    """
    Get all UUIDs from the src_metadatas.json file.
    
    Args:
        src_metadatas_path: Path to the src_metadatas.json file
        
    Returns:
        List of all UUIDs found in the file
    """
    try:
        with open(src_metadatas_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        
        filings_by_form = data.get("filings_by_form", {})
        uuids = []
        
        for form_type, filings in filings_by_form.items():
            for filing in filings:
                uuid_val = filing.get("uuid")
                if uuid_val:
                    uuids.append(uuid_val)
        
        return uuids
    except Exception as e:
        print(f"Error loading src_metadatas.json: {e}")
        return []


def load_earnings_calendar(ticker: str) -> Optional[Dict[str, Any]]:
    """
    Load earnings calendar data for a given ticker.
    
    Args:
        ticker: Stock ticker symbol
        
    Returns:
        Dictionary containing earnings calendar data or None if not found
    """
    try:
        earnings_cal_path = os.path.join(DATA_ROOT, "lkhs", "bronze", "e_cal", f"{ticker}_earnings_cal.json")
        if os.path.exists(earnings_cal_path):
            with open(earnings_cal_path, "r", encoding="utf-8") as f:
                earnings_cal = json.load(f)
            print(f"Loaded earnings calendar for {ticker}")
            return earnings_cal
        else:
            print(f"No earnings calendar found for {ticker} at {earnings_cal_path}")
            return None
    except Exception as e:
        print(f"Error loading earnings calendar for {ticker}: {e}")
        return None


def init_parser(pdf_path: str, earnings_cal: Optional[Dict[str, Any]] = None, source_catalog: str = None):
    """
    Initializes and runs the PDF parser for a given file.
    Enhanced with thread safety and better error handling.
    Returns both the DocumentMetadata object and the UUID for document identification.
    For existing files, loads the current src_metadata.json to update src_metadatas.json.
    
    Args:
        pdf_path: Path to the PDF file
        earnings_cal: Optional earnings calendar data for fiscal quarter determination
    """
    pdf_path_without_affix = pdf_path.rsplit(".", 1)[0]
    src_metadata_path = pdf_path_without_affix + "/src_metadata.json"
    
    if os.path.exists(src_metadata_path):
        print(f"Loading existing src_metadata.json for {pdf_path}")
        # Load existing metadata to get current information
        try:
            with open(src_metadata_path, "r", encoding="utf-8") as f:
                existing_metadata = json.load(f)
            
            # Extract the src_metadata (could be dict or DocumentMetadata object)
            src_metadata_dict = existing_metadata.get("src_metadata", {})
            uuid_val = existing_metadata.get("uuid")
            
            # Convert dict to DocumentMetadata object if needed
            if isinstance(src_metadata_dict, dict):
                try:
                    from document_metadata import DocumentMetadata
                    src_metadata_obj = DocumentMetadata(**src_metadata_dict)
                except Exception as e:
                    print(f"Warning: Could not convert dict to DocumentMetadata: {e}")
                    # Create a mock object with model_dump method for compatibility
                    class MockDocumentMetadata:
                        def __init__(self, data):
                            self.data = data
                        def model_dump(self):
                            return self.data
                    src_metadata_obj = MockDocumentMetadata(src_metadata_dict)
            else:
                src_metadata_obj = src_metadata_dict
            
            print(f"Successfully loaded existing metadata for {Path(pdf_path).name} (UUID: {uuid_val})")
            return src_metadata_obj, uuid_val
        except Exception as e:
            print(f"Warning: Could not load existing src_metadata.json: {e}")
            return None, None
    
    print(f"Starting parsing for: {pdf_path}")
    parser = PDFParserFitz(pdf_path, earnings_cal, source_catalog)
    summary = parser.parse_document()
    print(f"Successfully parsed {Path(pdf_path).name}.")
    print(f"Summary: {summary}")
    print(f"Outputs saved to: {parser.output_dir}")
    
    # Load the generated metadata to get UUID
    try:
        with open(parser.output_dir / "src_metadata.json", "r", encoding="utf-8") as f:
            generated_metadata = json.load(f)
        return parser.src_metadata, generated_metadata.get("uuid")
    except Exception as e:
        print(f"Warning: Could not load generated src_metadata.json: {e}")
        return parser.src_metadata, None

if __name__ == "__main__":
    if not DATA_ROOT:
        raise ValueError("DATA_ROOT environment variable is not set. Please set it in your .env.local file.")

    TARGET_PATH = os.path.join(DATA_ROOT, "lkhs", "bronze", "src")

    if not os.path.exists(TARGET_PATH):
        print(f"Target path not found: {TARGET_PATH}")
        exit()

    for ticker_subfolder in TICKER_LIST:
        
        ticker_path = os.path.join(TARGET_PATH, ticker_subfolder)
        if not os.path.isdir(ticker_path):
            continue

        # Load earnings calendar data for this ticker
        earnings_cal = load_earnings_calendar(ticker_subfolder)
        print(f"Loaded earnings calendar for {ticker_subfolder}: {earnings_cal}")
        
        all_metadata_for_ticker = []
        
        for source_catalog in SOURCE_CATALOG_LIST:
            # Look for research reports
            SOURCE_PATH = os.path.join(ticker_path, source_catalog)
            
            if not os.path.exists(SOURCE_PATH) or not os.path.isdir(SOURCE_PATH):
                print(f"No '{source_catalog}' folder in {ticker_subfolder}, skipping.")
                continue
            else:
                print(f"\nProcessing directory: {SOURCE_PATH}")

            for filename in os.listdir(SOURCE_PATH):
                if filename.lower().endswith(".pdf"):
                    pdf_path = os.path.join(SOURCE_PATH, filename)
                    
                    print(f"  Parsing PDF: {filename}")
                    try:
                        if source_catalog in [
                            # "earnings_call_transcript",
                            "company_presentation",
                            "company_presentation_slides",
                            "earnings_slides",
                        ]:
                            metadata, doc_uuid = init_parser(pdf_path, earnings_cal, source_catalog)
                        else:
                            metadata, doc_uuid = init_parser(pdf_path, earnings_cal)

                        if metadata:
                            try:
                                metadata_dict = metadata.model_dump()
                                metadata_dict["filename"] = filename
                                metadata_dict["filepath"] = str(Path(pdf_path).relative_to(DATA_ROOT))
                                metadata_dict["uuid"] = doc_uuid  # Add UUID for document identification
                                all_metadata_for_ticker.append(metadata_dict)
                            except AttributeError as e:
                                print(f"  ERROR: Metadata object doesn't have model_dump method: {e}")
                                print(f"  Metadata type: {type(metadata)}")
                            except Exception as e:
                                print(f"  ERROR processing metadata for {filename}: {e}")

                    except Exception as e:
                        print(f"  ERROR processing {filename}: {e}")

        # Save aggregated metadata for the ticker
        if all_metadata_for_ticker:
            output_json_path = os.path.join(ticker_path, "src_metadatas.json")
            
            # 1. Check if src_metadatas.json already exists
            old_structure = {}
            if os.path.exists(output_json_path):
                try:
                    with open(output_json_path, "r", encoding="utf-8") as f:
                        old_structure = json.load(f)
                    print(f"Found existing src_metadatas.json with {old_structure.get('total_filings', 0)} filings")
                except Exception as e:
                    print(f"Warning: Could not load existing src_metadatas.json: {e}")
                    old_structure = {}
            
            # Group new metadata by form_type
            new_filings_by_form = {}
            for meta in all_metadata_for_ticker:
                form_type = meta.get("form_type", "unknown")
                if form_type not in new_filings_by_form:
                    new_filings_by_form[form_type] = []
                new_filings_by_form[form_type].append(meta)

            # 3. Merge old_structure + all_metadata_for_ticker
            # Start with existing filings_by_form or empty dict
            merged_filings_by_form = old_structure.get("filings_by_form", {}).copy()
            
            # Calculate the next available citation_id counter
            # Collect all existing citation_ids to find the maximum
            max_citation_num = 0
            for form_type, filings in merged_filings_by_form.items():
                for filing in filings:
                    citation_id = filing.get("citation_id", "")
                    if citation_id and citation_id.startswith("ref"):
                        try:
                            num = int(citation_id[3:])  # Extract number from "refX"
                            max_citation_num = max(max_citation_num, num)
                        except (ValueError, IndexError):
                            pass
            
            # Start counter for new documents
            citation_counter = max_citation_num
            
            # Add/merge new filings by form_type
            for form_type, new_filings in new_filings_by_form.items():
                if form_type in merged_filings_by_form:
                    # Create mappings for existing filings
                    existing_by_uuid = {filing.get("uuid"): filing for filing in merged_filings_by_form[form_type] if filing.get("uuid")}
                    existing_by_filename = {filing.get("filename"): filing for filing in merged_filings_by_form[form_type]}
                    
                    # Process each new filing
                    for new_filing in new_filings:
                        new_uuid = new_filing.get("uuid")
                        new_filename = new_filing.get("filename")
                        
                        # Check by UUID first (primary), then by filename (fallback)
                        if new_uuid and new_uuid in existing_by_uuid:
                            # Update existing entry with current src_metadata.json data
                            existing_filing = existing_by_uuid[new_uuid]
                            print(f"  Updating existing entry (UUID match): {new_filename} (UUID: {new_uuid})")
                            
                            # Update the existing entry with new data, but preserve citation_id
                            existing_citation_id = existing_filing.get("citation_id")
                            existing_filing.update(new_filing)
                            if existing_citation_id:
                                existing_filing["citation_id"] = existing_citation_id
                            
                        elif new_filename in existing_by_filename:
                            # Update existing entry by filename (fallback for files without UUID)
                            existing_filing = existing_by_filename[new_filename]
                            print(f"  Updating existing entry (filename match): {new_filename}")
                            
                            # Update the existing entry with new data, but preserve citation_id
                            existing_citation_id = existing_filing.get("citation_id")
                            existing_filing.update(new_filing)
                            if existing_citation_id:
                                existing_filing["citation_id"] = existing_citation_id
                            
                        else:
                            # Add new filing with new citation_id
                            citation_counter += 1
                            new_filing["citation_id"] = f"ref{citation_counter}"
                            merged_filings_by_form[form_type].append(new_filing)
                            print(f"  Added new filing: {new_filename} (citation_id: ref{citation_counter})")
                else:
                    # New form_type, add all filings with new citation_ids
                    for new_filing in new_filings:
                        citation_counter += 1
                        new_filing["citation_id"] = f"ref{citation_counter}"
                    merged_filings_by_form[form_type] = new_filings
                    print(f"  Added new form type '{form_type}' with {len(new_filings)} filings")

            # Calculate total filings across all form types
            total_filings = sum(len(filings) for filings in merged_filings_by_form.values())

            final_structure = {
                "ticker": ticker_subfolder,
                "last_updated": datetime.now().strftime("%Y%m%d_%H%M%S"),
                "total_filings": total_filings,
                "filings_by_form": merged_filings_by_form,
                "summary": {k: len(v) for k, v in merged_filings_by_form.items()}
            }
            
            with open(output_json_path, "w", encoding="utf-8") as f:
                # Use default=str to handle potential non-serializable types like datetime
                json.dump(final_structure, f, indent=4, ensure_ascii=False, default=str)
            
            print(f"Saved merged source metadata to: {output_json_path}")
            print(f"  Total filings: {total_filings} (processed {len(all_metadata_for_ticker)} files)")
            print(f"  Form types: {list(final_structure['summary'].keys())}")
            
            # Count documents with UUIDs
            documents_with_uuid = sum(1 for filings in merged_filings_by_form.values() 
                                    for filing in filings if filing.get("uuid"))
            print(f"  Documents with UUID: {documents_with_uuid}/{total_filings}")
            
            # Count updated vs new documents
            updated_count = 0
            new_count = 0
            for form_type, filings in merged_filings_by_form.items():
                for filing in filings:
                    if filing.get("uuid") and any(f.get("uuid") == filing.get("uuid") for f in all_metadata_for_ticker):
                        updated_count += 1
                    else:
                        new_count += 1
            
            print(f"  Updated existing entries: {updated_count}")
            print(f"  Added new entries: {new_count}")
            




